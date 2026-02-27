#!/usr/bin/env python3
"""
Import leads from Exported Deals.xlsx + Exported People.xlsx into Supabase.
Filters to Prereads US pipeline only.

Usage:
  python scripts/import_leads.py \
    --deals "../Exported Deals.xlsx" \
    --people "../Exported People.xlsx"
"""

import sys
import os
import argparse
import openpyxl
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.config import get_settings
from app.database import get_db
from app.services.pipedrive import upsert_companies_from_xlsx, upsert_contacts_from_xlsx, extract_domain
from rich.console import Console
from rich.progress import track

console = Console()
PIPELINE = "Prereads US"


def load_xlsx(path: str) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        rows.append(row)
    return headers, rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--deals", default="Exported Deals.xlsx")
    parser.add_argument("--people", default="Exported People.xlsx")
    args = parser.parse_args()

    console.print(f"[bold blue]xAID Signal Platform — Lead Importer[/bold blue]")
    console.print(f"Loading deals from: {args.deals}")
    _, deals_rows = load_xlsx(args.deals)

    # Filter to Prereads US
    prereads = [r for r in deals_rows if r.get("Deal - Pipeline") == PIPELINE]
    console.print(f"[green]Found {len(prereads)} deals in '{PIPELINE}' pipeline[/green]")

    console.print(f"Loading people from: {args.people}")
    _, people_rows = load_xlsx(args.people)
    console.print(f"[green]Found {len(people_rows)} contacts[/green]")

    db = get_db()

    # Upsert companies
    console.print("\n[bold]Importing companies...[/bold]")
    company_id_map = upsert_companies_from_xlsx(prereads, db)
    console.print(f"[green]Upserted {len(company_id_map)} companies[/green]")

    # Try to fill domain from org name (basic extraction)
    console.print("Enriching company domains...")
    for org_id, uuid in company_id_map.items():
        # Look up the org row to get website if available
        for row in prereads:
            if row.get("Deal - Organization ID") == org_id:
                org_name = row.get("Deal - Organization", "")
                # Try to guess domain from org name (very basic)
                # Real enrichment would use Clearbit/Apollo
                break

    # Upsert contacts
    console.print("\n[bold]Importing contacts...[/bold]")
    count = upsert_contacts_from_xlsx(people_rows, company_id_map, db)
    console.print(f"[green]Upserted {count} contacts[/green]")

    console.print("\n[bold green]✓ Import complete![/bold green]")
    console.print(f"  Companies: {len(company_id_map)}")
    console.print(f"  Contacts:  {count}")


if __name__ == "__main__":
    main()
